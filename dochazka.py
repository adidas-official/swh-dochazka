from os import system
import datetime
from pathlib import Path
import re
import logging
import months_cz
import openpyxl

import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename

logging.basicConfig(level=logging.DEBUG, filename='log/error.log', format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
staff_data = {}


def choose_month():
    file = askopenfilename(parent=root, title='Vyberte soubor', filetypes=[('text file', '*.txt')])
    return Path(file).resolve()


# prepare input data
def get_content_data():
    content = ''
    with open(choose_month(), 'r') as a_file:
        line = True
        while line:
            line = a_file.readline()
            content += line.replace('  ', '')
    return content


# split chunks of data for each person
# return array of people
def split_to_chunks(content):

    delimiter_re = re.compile(r'--+')
    delimiter_mo = delimiter_re.findall(content)
    delimiter = '-' * 80
    if delimiter_mo:
        delimiter = delimiter_mo[0]

    staff = content.split(delimiter)
    staff_members = []
    for i in range(1, len(staff), 2):
        staff_member = staff[i-1] + staff[i]
        member_data = staff_member.split('\n')
        while '' in member_data:
            member_data.remove('')
        if member_data[0].startswith('Odpraco') or member_data[0].startswith('Nepří'):
            member_data.remove(member_data[0])
        staff_members.append(member_data)
    return staff_members


def make_sheet():
    c = get_content_data()
    all_staff = split_to_chunks(c)

    name_num_re = re.compile(r'(\d{4,})+"\s/\s([\w\sěščřžýáíéúůĚŠČŘŽÝÁÍÉÚŮóÓ,-.]+)')
    date_re = re.compile(r'(^\d{2}\.\d{2}\.\d{2})\s(\d{2}:\d{2}).*(\d{2}:\d{2})\s*\S*\s([A-Z])?.*(\w{2})')

    month_re = re.compile(r'\d{2}\.\d{2}\.\d{4}')
    month_mo = month_re.search(c)
    if month_mo:
        time_period = month_mo.group().split('.')
        month_num = int(time_period[1]) - 1
        year = time_period[2]
        month = months_cz.months_cz[month_num]
    else:
        year = datetime.datetime.now().year
        month = 'mesic'

    member_id = False
    for person in all_staff:
        for data in person:
            name_id_mo = name_num_re.findall(data)
            if name_id_mo:
                member_id, member_name = name_id_mo[0]
                staff_data.setdefault(member_id, {
                        'name': member_name,
                        'attendance': []
                    }
                )

            att_mo = date_re.findall(data)
            if att_mo and member_id:
                # logging.info(att_mo[0]) # ('03.01.22', '07:00', '15:00')
                staff_data[member_id]['attendance'].append(att_mo[0])

    template = Path('inc/template.xlsx')
    if template.exists():
        wb = openpyxl.load_workbook(template)
        ws = wb.active

        for k, v in staff_data.items():
            new_worksheet = wb.copy_worksheet(ws)
            new_worksheet.title = v['name']
            new_worksheet.cell(1, 3).value = v['name']
            new_worksheet.cell(2, 3).value = f'{month} {year}'
            new_worksheet.cell(3, 3).value = k

            for workday in v['attendance']:

                row = int(workday[0][:2]) + 5
                new_worksheet.cell(row, 2).value = workday[4]
                new_worksheet.cell(row, 3).value = workday[1]
                new_worksheet.cell(row, 3).number_format = 'H:mm;@'
                new_worksheet.cell(row, 4).value = workday[2]
                new_worksheet.cell(row, 4).number_format = 'H:mm;@'
                new_worksheet.cell(row, 5).value = f'=(D{row}-C{row})*24'
                new_worksheet.cell(row, 6).value = workday[3]

        wb.remove(ws)
        save_location = Path.home() / 'Desktop'
        if save_location.exists():
            wb.save(f'{save_location}/{month} {year}.xlsx')
            system(f'start {save_location}')
        else:
            logging.error(f'Slozka {save_location} neexistuje')

        root.destroy()


root = tk.Tk()
root.iconbitmap('inc/swh.ico')


def center_window(window, w, h):
    screen_w = window.winfo_screenwidth()
    screen_h = window.winfo_screenheight()
    left_point = int(screen_w / 2 - w / 2)
    top_point = int(screen_h / 2 - h / 2)

    root.geometry(f'{w}x{h}+{left_point}+{top_point}')


center_window(root, 500, 400)

root.configure(
    background='white'
)
root.title('Dochazka')
root.attributes('-alpha', 0.9)

instructions = ttk.Label(
    root,
    text='Vyberte txt soubor exportovaný z ucto2022\nSoubor bude uložen na plochu',
    background='white',
    anchor='center',
    justify=tk.CENTER,
    font=('Helvetica', 13)
)


instructions.pack(
    expand=True
)

btn_text = tk.StringVar()
btn = ttk.Button(
    root,
    command=lambda: make_sheet(),
    textvariable=btn_text,
    padding=15
)

btn_text.set('Procházet')
btn.pack(
    expand=True
)

root.mainloop()

